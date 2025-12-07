#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script to create a sample Excel file for testing client import functionality.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from django.contrib.auth import get_user_model
from clients.models import ClientGroup
import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

User = get_user_model()

def create_sample_excel():
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Define headers that match our model fields
    # Note: We need to use the field names as expected by import_export
    headers = [
        'name',  # клиент (обязательное поле)
        'employee__username',  # сотрудник (необязательное поле)
        'client_groups__name',  # группа (необязательное поле)
        'trading_point_name',  # название торговой точки (необязательное поле)
        'trading_point_address'  # адрес торговой точки (необязательное поле)
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add sample data
    sample_data = [
        ["ООО Ромашка", "admin", "VIP", "Магазин №1", "Москва, ул. Ленина, д.10"],
        ["ООО Глобус", "moderator", "Regular", "Супермаркет", "СПб, пр. Невский, д.5"],
        ["ИП Петров", "employee1", "", "Продукты", "Екатеринбург, ул. Пушкина, д.3"],
        ["ООО Свет", "", "Premium", "Торговый центр", "Новосибирск, пл. Центральная, д.1"],
        ["ООО Луч", "employee2", "", "", "Казань, ул. Гагарина, д.15"]
    ]
    
    # Add sample data to the worksheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    filename = "/workspace/sample_clients_import.xlsx"
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    print("Headers:", headers)
    print("Note: employee__username should match existing usernames")
    print("Note: client_groups__name should match existing group names")

if __name__ == "__main__":
    create_sample_excel()