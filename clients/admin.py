# -*- coding: utf-8 -*-
"""
Created on Fri Nov 28 20:58:43 2025

@author: Professional
"""

"""
Admin interface for client management.

This module defines the admin interface configuration for client models.
Provides Russian language interface and inline editing capabilities.
"""

from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse
from django.contrib import messages
from .models import Client, ClientGroup
import openpyxl
from openpyxl import load_workbook


class ClientInline(admin.TabularInline):
    """Inline client editing for client groups."""
    model = Client.client_groups.through
    extra = 1
    verbose_name = _('Клиент')
    verbose_name_plural = _('Клиенты')


@admin.register(ClientGroup)
class ClientGroupAdmin(admin.ModelAdmin):
    """
    Admin interface for ClientGroup model.
    
    Provides inline editing of clients within groups.
    """
    
    inlines = [ClientInline]
    list_display = ('name', 'get_client_count', 'created_at')
    search_fields = ('name',)
    list_per_page = 20
    
    def get_client_count(self, obj):
        """Return number of clients in group."""
        return obj.client_set.count()
    get_client_count.short_description = _('Количество клиентов')
    
    class Meta:
        verbose_name = _('Группа клиентов')
        verbose_name_plural = _('Группы клиентов')


@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    """
    Admin interface for Client model.
    
    Provides comprehensive client management interface.
    """

    list_display = ('name', 'employee', 'address', 'get_groups', 'created_at')
    list_filter = ('client_groups', 'employee')
    search_fields = ('name', 'address')
    filter_horizontal = ('client_groups',)
    list_per_page = 20
    
    def get_groups(self, obj):
        """Return comma-separated list of client groups."""
        return ', '.join([group.name for group in obj.client_groups.all()])
    get_groups.short_description = _('Группы')
    
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='client_import_excel'),
        ]
        return custom_urls + urls
    
    def import_excel(self, request):
        """Import clients from Excel file."""
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, _('Файл не выбран'))
                return HttpResponseRedirect(request.path_info)
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, _('Неверный формат файла. Пожалуйста, загрузите файл Excel (.xlsx или .xls)'))
                return HttpResponseRedirect(request.path_info)
            
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # Read headers from first row
                headers = []
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header:
                        headers.append(header.strip().lower())
                    else:
                        headers.append(f'column_{col}')
                
                # Expected headers
                client_col = None
                employee_col = None
                group_col = None
                shop_name_col = None
                shop_address_col = None
                
                for i, header in enumerate(headers):
                    if 'клиент' in header:
                        client_col = i
                    elif 'сотрудник' in header:
                        employee_col = i
                    elif 'групп' in header:
                        group_col = i
                    elif 'торгов' in header and 'точк' in header:
                        shop_name_col = i
                    elif 'адрес' in header and ('торгов' in header or 'точк' in header):
                        shop_address_col = i
                
                if client_col is None:
                    messages.error(request, _('Обязательное поле "клиент" не найдено в файле'))
                    return HttpResponseRedirect(request.path_info)
                
                created_count = 0
                updated_count = 0
                
                for row in range(2, ws.max_row + 1):  # Skip header row
                    client_name = ws.cell(row=row, column=client_col + 1).value
                    if not client_name:
                        continue
                    
                    client_name = str(client_name).strip()
                    if not client_name:
                        continue
                    
                    # Get or create client
                    client, created = Client.objects.get_or_create(name=client_name)
                    
                    # Update employee if provided
                    if employee_col is not None:
                        employee_name = ws.cell(row=row, column=employee_col + 1).value
                        if employee_name:
                            from users.models import CustomUser
                            try:
                                employee = CustomUser.objects.get(username=str(employee_name).strip())
                                client.employee = employee
                            except CustomUser.DoesNotExist:
                                messages.warning(request, f'Сотрудник "{employee_name}" не найден для клиента "{client_name}"')
                    
                    # Update address if provided
                    if shop_address_col is not None:
                        shop_address = ws.cell(row=row, column=shop_address_col + 1).value
                        if shop_address:
                            client.address = str(shop_address).strip()
                    
                    client.save()
                    
                    # Add to group if provided
                    if group_col is not None:
                        group_name = ws.cell(row=row, column=group_col + 1).value
                        if group_name:
                            group_name = str(group_name).strip()
                            client_group, group_created = ClientGroup.objects.get_or_create(name=group_name)
                            client.client_groups.add(client_group)
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                
                messages.success(request, f'Импорт завершен: создано {created_count} клиентов, обновлено {updated_count} клиентов')
                return HttpResponseRedirect(reverse('admin:clients_client_changelist'))
                
            except Exception as e:
                messages.error(request, f'Ошибка при импорте файла: {str(e)}')
                return HttpResponseRedirect(request.path_info)
        
        context = {
            'title': 'Импорт клиентов из Excel',
            'opts': self.model._meta,
            'has_permission': True,
        }
        return render(request, 'admin/clients/import_excel.html', context)
    
    class Meta:
        verbose_name = _('Клиент')
        verbose_name_plural = _('Клиенты')